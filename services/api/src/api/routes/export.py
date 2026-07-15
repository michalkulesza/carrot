import calendar as cal
import io
import uuid
from datetime import date as DateType, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import and_
from sqlalchemy.ext.asyncio import AsyncSession

from api.database import get_async_session
from api.models import MealPlanEntry
from api.routes.context import get_active_household_id
from api.users import User, current_active_user

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

router = APIRouter(prefix="/export", tags=["export"])

_DAY_HEADERS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
_GREEN = "356854"
_ROW_COUNT = 6


def _parse_month(month: str) -> tuple[int, int]:
    try:
        parts = month.split("-")
        return int(parts[0]), int(parts[1])
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid month format, use YYYY-MM")


async def _fetch_entries(
    year: int,
    month: int,
    user: User,
    household_id: uuid.UUID | None,
    session: AsyncSession,
) -> dict[str, str]:
    last_day = cal.monthrange(year, month)[1]
    start = DateType(year, month, 1)
    end = DateType(year, month, last_day)

    if household_id is not None:
        where = and_(
            MealPlanEntry.household_id == household_id,
            MealPlanEntry.date >= start,
            MealPlanEntry.date <= end,
        )
    else:
        where = and_(
            MealPlanEntry.user_id == user.id,
            MealPlanEntry.household_id.is_(None),
            MealPlanEntry.date >= start,
            MealPlanEntry.date <= end,
        )

    from sqlalchemy import select

    result = await session.execute(select(MealPlanEntry).where(where))
    return {str(e.date): e.recipe.title if e.recipe is not None else e.text or "" for e in result.scalars().all()}


def _week_mondays(year: int, month: int) -> list[DateType]:
    first = DateType(year, month, 1)
    last = DateType(year, month, cal.monthrange(year, month)[1])
    start = first - timedelta(days=first.weekday())
    weeks: list[DateType] = []
    d = start
    while d <= last:
        weeks.append(d)
        d += timedelta(days=7)
    return weeks


def _build_xlsx(by_date: dict[str, str], year: int, month: int) -> bytes:
    weeks = _week_mondays(year, month)
    total_rows = 1 + _ROW_COUNT
    edge = Side(style="thin", color=_GREEN)
    no_edge = Side(style=None)

    def _border(row_idx: int, col_idx: int) -> Border:
        return Border(
            top=edge if row_idx == 1 else no_edge,
            bottom=edge if row_idx == total_rows else no_edge,
            left=edge if col_idx == 1 else no_edge,
            right=edge if col_idx == 7 else no_edge,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Week Meal Planner"

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 24.24

    ws.append(_DAY_HEADERS)
    ws.row_dimensions[1].height = 31.22
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill("solid", fgColor=_GREEN)
        cell.font = Font(name="Times New Roman", size=16, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border(1, col)

    for wi in range(_ROW_COUNT):
        row_data: list[str | None] = []
        monday = weeks[wi] if wi < len(weeks) else None
        for i in range(7):
            if monday:
                ds = (monday + timedelta(days=i)).strftime("%Y-%m-%d")
                row_data.append(by_date.get(ds))
            else:
                row_data.append(None)
        ws.append(row_data)
        row_idx = wi + 2
        ws.row_dimensions[row_idx].height = 71.38
        bg = "FFFFFF" if wi % 2 == 0 else "F6F8F9"
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Roboto", size=14, color="434343")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _border(row_idx, col)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_pdf(by_date: dict[str, str], year: int, month: int) -> bytes:
    weeks = _week_mondays(year, month)

    header_style = ParagraphStyle(
        "header",
        fontName="Times-Roman",
        fontSize=12,
        textColor=colors.white,
        alignment=TA_CENTER,
        leading=14,
    )
    cell_style = ParagraphStyle(
        "cell",
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor("#434343"),
        alignment=TA_CENTER,
        leading=11,
    )

    data = [[Paragraph(h, header_style) for h in _DAY_HEADERS]]
    for wi in range(_ROW_COUNT):
        row = []
        monday = weeks[wi] if wi < len(weeks) else None
        for i in range(7):
            if monday:
                ds = (monday + timedelta(days=i)).strftime("%Y-%m-%d")
                text = by_date.get(ds, "")
            else:
                text = ""
            row.append(Paragraph(text, cell_style))
        data.append(row)

    page_w, _ = A4
    col_w = page_w / 7
    t = Table(
        data,
        colWidths=[col_w] * 7,
        rowHeights=[22] + [50] * _ROW_COUNT,
    )

    green = colors.HexColor("#356854")
    style_cmds = [
        ("BACKGROUND", (0, 0), (6, 0), green),
        ("BOX", (0, 0), (-1, -1), 1, green),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for wi in range(_ROW_COUNT):
        if wi % 2 == 1:
            style_cmds.append(("BACKGROUND", (0, wi + 1), (6, wi + 1), colors.HexColor("#F6F8F9")))

    t.setStyle(TableStyle(style_cmds))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=0,
        rightMargin=0,
        topMargin=0,
        bottomMargin=0,
    )
    doc.build([t])
    return buf.getvalue()


@router.get("/meal-plan.xlsx")
async def export_meal_plan_xlsx(
    month: str,
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(get_async_session),
    household_id: uuid.UUID | None = Depends(get_active_household_id),
) -> Response:
    year, m = _parse_month(month)
    by_date = await _fetch_entries(year, m, user, household_id, session)
    data = _build_xlsx(by_date, year, m)
    filename = f"meal-plan-{year}-{m:02d}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/meal-plan.pdf")
async def export_meal_plan_pdf(
    month: str,
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(get_async_session),
    household_id: uuid.UUID | None = Depends(get_active_household_id),
) -> Response:
    year, m = _parse_month(month)
    by_date = await _fetch_entries(year, m, user, household_id, session)
    data = _build_pdf(by_date, year, m)
    filename = f"meal-plan-{year}-{m:02d}.pdf"
    return Response(
        content=data,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
